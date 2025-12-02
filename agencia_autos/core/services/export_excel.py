"""
Servicio para exportar reportes a Excel con formato profesional
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from datetime import datetime


def exportar_disponibilidad_excel(datos, filtros=None):
    """
    Exporta el reporte de disponibilidad a Excel con formato profesional
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Disponibilidad"
    
    # Estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título del reporte
    ws.merge_cells('A1:D1')
    ws['A1'] = 'REPORTE DE DISPONIBILIDAD DE VEHÍCULOS'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Fecha de generación
    ws.merge_cells('A2:D2')
    ws['A2'] = f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Filtros aplicados
    if filtros:
        row = 3
        if filtros.get('fecha_desde') or filtros.get('fecha_hasta'):
            ws.merge_cells(f'A{row}:D{row}')
            fecha_texto = f"Período: {filtros.get('fecha_desde', 'Inicio')} - {filtros.get('fecha_hasta', 'Hoy')}"
            ws[f'A{row}'] = fecha_texto
            row += 1
        if filtros.get('marca'):
            ws.merge_cells(f'A{row}:D{row}')
            ws[f'A{row}'] = f"Marca: {filtros['marca']}"
            row += 1
        if filtros.get('tipo'):
            ws.merge_cells(f'A{row}:D{row}')
            ws[f'A{row}'] = f"Tipo: {filtros['tipo']}"
            row += 1
        start_row = row + 1
    else:
        start_row = 4
    
    # Encabezados
    headers = ['Marca', 'Tipo de Vehículo', 'Cantidad Disponible', 'Precio Promedio']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Datos
    for row_num, dato in enumerate(datos, start_row + 1):
        ws.cell(row=row_num, column=1, value=dato['marca']).border = border
        ws.cell(row=row_num, column=2, value=dato['tipo_vehiculo']).border = border
        ws.cell(row=row_num, column=3, value=dato['cantidad_disponible']).border = border
        ws.cell(row=row_num, column=3).alignment = Alignment(horizontal='center')
        
        precio_cell = ws.cell(row=row_num, column=4, value=float(dato['precio_promedio']))
        precio_cell.number_format = '$#,##0.00'
        precio_cell.border = border
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    
    # Totales
    if datos:
        total_row = len(datos) + start_row + 1
        ws.merge_cells(f'A{total_row}:B{total_row}')
        ws[f'A{total_row}'] = 'TOTAL'
        ws[f'A{total_row}'].font = Font(bold=True)
        ws[f'A{total_row}'].alignment = Alignment(horizontal='right')
        
        total_cantidad = sum(d['cantidad_disponible'] for d in datos)
        ws[f'C{total_row}'] = total_cantidad
        ws[f'C{total_row}'].font = Font(bold=True)
        ws[f'C{total_row}'].alignment = Alignment(horizontal='center')
        ws[f'C{total_row}'].border = border
        
        promedio_precio = sum(float(d['precio_promedio']) for d in datos) / len(datos)
        ws[f'D{total_row}'] = promedio_precio
        ws[f'D{total_row}'].number_format = '$#,##0.00'
        ws[f'D{total_row}'].font = Font(bold=True)
        ws[f'D{total_row}'].border = border
    
    # Preparar respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'reporte_disponibilidad_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


def exportar_ventas_excel(datos, filtros=None):
    """
    Exporta el reporte de ventas a Excel con formato profesional
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas"
    
    # Estilos
    header_fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws.merge_cells('A1:F1')
    ws['A1'] = 'REPORTE DE VENTAS'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Fecha de generación
    ws.merge_cells('A2:F2')
    ws['A2'] = f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Filtros aplicados
    start_row = 4
    if filtros:
        row = 3
        if filtros.get('fecha_desde') or filtros.get('fecha_hasta'):
            ws.merge_cells(f'A{row}:F{row}')
            fecha_texto = f"Período: {filtros.get('fecha_desde', 'Inicio')} - {filtros.get('fecha_hasta', 'Hoy')}"
            ws[f'A{row}'] = fecha_texto
            row += 1
        start_row = row + 1
    
    # Encabezados
    headers = ['ID Venta', 'Fecha', 'Cliente', 'Empleado', 'Total', 'Estado']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Datos
    for row_num, venta in enumerate(datos, start_row + 1):
        ws.cell(row=row_num, column=1, value=venta.id).border = border
        ws.cell(row=row_num, column=2, value=venta.fecha_venta.strftime('%d/%m/%Y')).border = border
        ws.cell(row=row_num, column=3, value=venta.cliente.nombre_completo).border = border
        ws.cell(row=row_num, column=4, value=venta.empleado.nombre_completo).border = border
        
        total_cell = ws.cell(row=row_num, column=5, value=float(venta.total_venta))
        total_cell.number_format = '$#,##0.00'
        total_cell.border = border
        
        ws.cell(row=row_num, column=6, value=venta.estado_venta).border = border
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    
    # Totales
    if datos:
        total_row = len(datos) + start_row + 1
        ws.merge_cells(f'A{total_row}:D{total_row}')
        ws[f'A{total_row}'] = 'TOTAL'
        ws[f'A{total_row}'].font = Font(bold=True)
        ws[f'A{total_row}'].alignment = Alignment(horizontal='right')
        
        total_ventas = sum(float(v.total_venta) for v in datos)
        ws[f'E{total_row}'] = total_ventas
        ws[f'E{total_row}'].number_format = '$#,##0.00'
        ws[f'E{total_row}'].font = Font(bold=True)
        ws[f'E{total_row}'].border = border
    
    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'reporte_ventas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response
